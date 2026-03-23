"""Import BLS data (QCEW county employment + occupational projections)
into pathway_data.db, mapped to the 7 pathway families.

Creates two tables:
  - bls_employment: QCEW quarterly employment by pathway family (4-county aggregate)
  - bls_projections: 2022-2032 occupational projections by pathway family

Usage:
    python scripts/import_bls_data.py
"""

import csv
import os
import sqlite3

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("WARNING: openpyxl not installed — projection import will be skipped")

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
BLS_DIR = os.path.join(BASE_DIR, "data", "bls")
TARGET_DB = os.path.join(BASE_DIR, "data", "pathway_data.db")

# NAICS 2-digit → pathway family mapping
NAICS_TO_PATHWAY = {
    "62": "healthcare",
    "52": "business",
    "53": "business",
    "54": "business",
    "55": "business",
    "31-33": "manufacturing",
    "23": "manufacturing",
    "42": "logistics",
    "44-45": "logistics",
    "48-49": "logistics",
    "51": "tech",
    "61": "education",
    "92": "law_public",
}

# Projection sheet occupation title keywords → pathway family
PROJECTION_KEYWORDS = {
    "healthcare": [
        "nurse", "health", "medical", "dental", "pharm", "therap",
        "physician", "surgeon", "diagnostic", "clinical", "care aide",
        "respiratory", "veterinar", "optom", "chiropract", "dietit",
        "radiolog", "sonograph", "psychiatric", "anesthesi",
    ],
    "business": [
        "accountant", "financial", "business", "marketing", "manager",
        "human resource", "budget", "credit", "loan", "insurance",
        "real estate", "management analyst", "logistician",
        "compensation", "payroll", "cost estimator",
    ],
    "manufacturing": [
        "machinist", "welder", "electrician", "plumber", "carpenter",
        "construction", "industrial", "mechanic", "hvac", "manufacturing",
        "assembler", "inspector", "fitter", "sheet metal", "millwright",
    ],
    "logistics": [
        "truck driver", "warehouse", "shipping", "freight", "supply chain",
        "stock", "material", "dispatch", "cargo", "forklift", "courier",
        "delivery", "transportation", "postal",
    ],
    "tech": [
        "software", "computer", "data", "web developer", "information security",
        "network", "database", "programmer", "systems admin", "it ",
        "artificial intelligence", "cloud",
    ],
    "education": [
        "teacher", "education", "tutor", "instructional", "librarian",
        "school", "postsecondary", "special education",
    ],
    "law_public": [
        "lawyer", "paralegal", "legal", "judge", "court", "probation",
        "social worker", "counselor", "community service",
    ],
}


def classify_occupation(title):
    """Map an occupation title to a pathway family using keyword matching."""
    title_lower = title.lower()
    for family, keywords in PROJECTION_KEYWORDS.items():
        for kw in keywords:
            if kw in title_lower:
                return family
    return None


def import_qcew(db):
    """Import QCEW county employment data, aggregated by pathway family."""
    print("\n1. Importing QCEW county employment data...")

    db.execute("DROP TABLE IF EXISTS bls_employment")
    db.execute("""
        CREATE TABLE bls_employment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pathway_family TEXT NOT NULL,
            naics_code TEXT,
            naics_title TEXT,
            county_fips TEXT,
            county_name TEXT,
            quarter INTEGER,
            establishments INTEGER DEFAULT 0,
            employment INTEGER DEFAULT 0,
            avg_weekly_wage INTEGER DEFAULT 0,
            total_quarterly_wages INTEGER DEFAULT 0,
            oty_employment_chg INTEGER DEFAULT 0,
            oty_employment_pct_chg REAL DEFAULT 0,
            oty_wage_chg INTEGER DEFAULT 0,
            oty_wage_pct_chg REAL DEFAULT 0
        )
    """)

    csv_files = [f for f in os.listdir(BLS_DIR) if f.endswith(".csv")]
    total_rows = 0

    for fname in sorted(csv_files):
        fpath = os.path.join(BLS_DIR, fname)
        county_name = fname.split(" ", 2)[-1].replace(".csv", "")
        print(f"   Processing {county_name}...")

        with open(fpath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                own_code = row.get("own_code", "")
                agglvl = row.get("agglvl_code", "")
                industry_code = row.get("industry_code", "")

                # Private sector, NAICS 2-digit only
                if own_code != "5" or agglvl not in ("74", "75"):
                    continue

                pathway = NAICS_TO_PATHWAY.get(industry_code)
                if not pathway:
                    continue

                def safe_int(val):
                    try:
                        return int(str(val).replace(",", "").replace("$", "").strip())
                    except (ValueError, TypeError):
                        return 0

                def safe_float(val):
                    try:
                        return float(str(val).replace(",", "").strip())
                    except (ValueError, TypeError):
                        return 0.0

                db.execute("""
                    INSERT INTO bls_employment (
                        pathway_family, naics_code, naics_title,
                        county_fips, county_name, quarter,
                        establishments, employment, avg_weekly_wage,
                        total_quarterly_wages,
                        oty_employment_chg, oty_employment_pct_chg,
                        oty_wage_chg, oty_wage_pct_chg
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    pathway,
                    industry_code,
                    row.get("industry_title", ""),
                    row.get("area_fips", ""),
                    county_name,
                    safe_int(row.get("qtr", "0")),
                    safe_int(row.get("qtrly_estabs_count", "0")),
                    safe_int(row.get("month3_emplvl", "0")),
                    safe_int(row.get("avg_wkly_wage", "0")),
                    safe_int(row.get("total_qtrly_wages", "0")),
                    safe_int(row.get("oty_month3_emplvl_chg", "0")),
                    safe_float(row.get("oty_month3_emplvl_pct_chg", "0")),
                    safe_int(row.get("oty_avg_wkly_wage_chg", "0")),
                    safe_float(row.get("oty_avg_wkly_wage_pct_chg", "0")),
                ))
                total_rows += 1

    db.execute("CREATE INDEX idx_bls_emp_family ON bls_employment(pathway_family)")
    db.execute("CREATE INDEX idx_bls_emp_qtr ON bls_employment(quarter)")
    print(f"   {total_rows} employment rows imported")
    return total_rows


def import_projections(db):
    """Import occupational projections from the analysis pack Excel."""
    if not openpyxl:
        print("\n2. SKIPPED — openpyxl not installed")
        return 0

    xlsx_path = os.path.join(BLS_DIR, "kc_projection_22_32_analysis_pack.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"\n2. SKIPPED — {xlsx_path} not found")
        return 0

    print("\n2. Importing occupational projections...")

    db.execute("DROP TABLE IF EXISTS bls_projections")
    db.execute("""
        CREATE TABLE bls_projections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pathway_family TEXT,
            title TEXT NOT NULL,
            now_next_later TEXT,
            employment_2022 INTEGER DEFAULT 0,
            employment_2032 INTEGER DEFAULT 0,
            net_change INTEGER DEFAULT 0,
            growth_pct REAL DEFAULT 0,
            median_wage INTEGER DEFAULT 0,
            mean_wage INTEGER DEFAULT 0,
            annual_openings INTEGER DEFAULT 0,
            education_required TEXT,
            grade TEXT,
            source_sheet TEXT
        )
    """)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    total_rows = 0

    # Import from key analysis sheets
    target_sheets = {
        "Top Openings": {
            "cols": ["title", "now_next_later", "employment_2022", "employment_2032",
                     "net_change", "growth_pct", "median_wage", "mean_wage",
                     "annual_openings", "exits", "transfers", "growth",
                     "education_required", "grade"],
        },
        "Top Net Growth": {
            "cols": ["title", "now_next_later", "employment_2022", "employment_2032",
                     "net_change", "growth_pct", "median_wage", "mean_wage",
                     "annual_openings", "education_required", "grade"],
        },
        "Fast Growth 1000+": {
            "cols": ["title", "now_next_later", "employment_2022", "employment_2032",
                     "net_change", "growth_pct", "median_wage", "mean_wage",
                     "annual_openings", "education_required", "grade"],
        },
        "High Opportunity": {
            "cols": ["title", "now_next_later", "employment_2022", "employment_2032",
                     "net_change", "growth_pct", "median_wage", "mean_wage",
                     "annual_openings", "education_required", "grade"],
            "skip_rows": 2,  # Has subtitle + blank row
        },
        "Projected Declines": {
            "cols": ["title", "now_next_later", "employment_2022", "employment_2032",
                     "net_change", "growth_pct", "median_wage", "annual_openings",
                     "education_required"],
        },
    }

    seen_titles = set()

    for sheet_name, config in target_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        skip = config.get("skip_rows", 0)
        col_map = config["cols"]

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= skip:  # skip header row(s)
                continue
            if not row or not row[0]:
                continue

            data = {}
            for j, col_name in enumerate(col_map):
                if j < len(row):
                    data[col_name] = row[j]

            title = str(data.get("title", ""))
            if not title or title in seen_titles:
                continue
            seen_titles.add(title)

            pathway = classify_occupation(title)

            def safe_int(val):
                try:
                    return int(float(val)) if val else 0
                except (ValueError, TypeError):
                    return 0

            def safe_float(val):
                try:
                    return float(val) if val else 0.0
                except (ValueError, TypeError):
                    return 0.0

            db.execute("""
                INSERT INTO bls_projections (
                    pathway_family, title, now_next_later,
                    employment_2022, employment_2032, net_change, growth_pct,
                    median_wage, mean_wage, annual_openings,
                    education_required, grade, source_sheet
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                pathway,
                title,
                str(data.get("now_next_later", "")),
                safe_int(data.get("employment_2022")),
                safe_int(data.get("employment_2032")),
                safe_int(data.get("net_change")),
                safe_float(data.get("growth_pct")),
                safe_int(data.get("median_wage")),
                safe_int(data.get("mean_wage", 0)),
                safe_int(data.get("annual_openings")),
                str(data.get("education_required", "")),
                str(data.get("grade", "")),
                sheet_name,
            ))
            total_rows += 1

    wb.close()
    db.execute("CREATE INDEX idx_bls_proj_family ON bls_projections(pathway_family)")
    print(f"   {total_rows} projection rows imported ({len(seen_titles)} unique occupations)")

    # Summary by pathway family
    cursor = db.execute("""
        SELECT pathway_family, COUNT(*), SUM(annual_openings), SUM(net_change)
        FROM bls_projections
        WHERE pathway_family IS NOT NULL
        GROUP BY pathway_family
        ORDER BY SUM(annual_openings) DESC
    """)
    print("\n   Pathway summary:")
    for row in cursor:
        print(f"     {row[0] or 'unmapped':<20} {row[1]:>3} occupations  "
              f"{row[2]:>6,} annual openings  {row[3]:>+6,} net growth")

    return total_rows


def import_pathway_summary(db):
    """Import the pathway-level summary data from the Pathway Summary sheet."""
    if not openpyxl:
        return 0

    xlsx_path = os.path.join(BLS_DIR, "kc_projection_22_32_analysis_pack.xlsx")
    if not os.path.exists(xlsx_path):
        return 0

    print("\n3. Importing pathway summary...")

    db.execute("DROP TABLE IF EXISTS bls_pathway_summary")
    db.execute("""
        CREATE TABLE bls_pathway_summary (
            pathway TEXT PRIMARY KEY,
            occupation_count INTEGER,
            employment_2022 INTEGER,
            employment_2032 INTEGER,
            net_change INTEGER,
            growth_rate REAL,
            annual_openings INTEGER,
            openings_share REAL,
            weighted_median_wage INTEGER,
            weighted_mean_wage INTEGER
        )
    """)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Pathway Summary"]
    total = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # skip title + blank
            continue
        if not row or not row[0] or row[0] == "Pathway":
            continue

        def safe_int(v):
            try: return int(float(v)) if v else 0
            except: return 0

        def safe_float(v):
            try: return float(v) if v else 0.0
            except: return 0.0

        db.execute("""
            INSERT OR REPLACE INTO bls_pathway_summary VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            str(row[0]).lower(),
            safe_int(row[1]),
            safe_int(row[2]),
            safe_int(row[3]),
            safe_int(row[4]),
            safe_float(row[5]),
            safe_int(row[6]),
            safe_float(row[7]),
            safe_int(row[8]),
            safe_int(row[9]),
        ))
        total += 1
        print(f"   {row[0]}: {safe_int(row[6]):,} annual openings, "
              f"{safe_float(row[5])*100:.1f}% growth")

    wb.close()
    print(f"   {total} pathway summary rows imported")
    return total


def main():
    if not os.path.exists(TARGET_DB):
        print(f"ERROR: {TARGET_DB} not found. Run export_pathway_data.py first.")
        return

    db = sqlite3.connect(TARGET_DB)

    emp_count = import_qcew(db)
    proj_count = import_projections(db)
    summary_count = import_pathway_summary(db)

    db.commit()

    print(f"\n{'=' * 60}")
    print(f"BLS import complete")
    print(f"  QCEW employment rows: {emp_count:,}")
    print(f"  Projection rows: {proj_count}")
    print(f"  Pathway summary rows: {summary_count}")
    print(f"  DB size: {os.path.getsize(TARGET_DB) / 1024:.0f} KB")
    print(f"{'=' * 60}")

    db.close()


if __name__ == "__main__":
    main()
